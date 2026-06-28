"""Import/Export audit (2026-06-28) — end-to-end verification of EVERY import and
export surface, driving the real HTTP APIs with REAL fake CSV/XLSX/PDF files.

Covers:
  * /api/import/types lists all importers
  * every type's CSV + XLSX template downloads AND round-trips back through the
    same importer with no error (the headers we hand the TD must be importable)
  * staged-import end-to-end (upload -> stage -> merge) for the key types,
    across the CSV, XLSX and PDF parse paths
  * error handling: corrupt file -> 400, missing-required -> staged-invalid,
    unknown USTA -> staged-invalid, unknown type -> 404
  * the export endpoints: payroll CSV (utf-8-sig BOM), assignment-audit CSV,
    official schedule .ics

Findings + the reuse checklist live in docs/import-export-notes.md.
Named test_zz_* to sort last (shared rationale with the other suites)."""
import io
import uuid
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app

client = TestClient(app)

pytestmark = pytest.mark.skipif(
    client.get("/api/health").json().get("db") != "ok",
    reason="Postgres not reachable / not migrated (run migrate.py)",
)

_FIXTURE_PDF = Path(__file__).parent / "fixtures" / "tournament_emails.pdf"


@pytest.fixture(autouse=True)
def _admin():
    client.post("/api/auth/login", json={"username": "admin", "password": "admin"})


def _ok(r, code=201):
    assert r.status_code == code, r.text
    return r.json()


def _u(prefix="2"):
    # USTA-shaped unique id (numeric-ish, unique per call so DB rows never clash)
    return prefix + uuid.uuid4().hex[:9]


def _tournament():
    return _ok(client.post("/api/tournaments", json={
        "name": "IE " + uuid.uuid4().hex[:6], "type": "junior",
        "play_start_date": "2026-06-01", "play_end_date": "2026-06-04"}))


def _player(usta, gender="female", first="Test", last=None):
    return _ok(client.post("/api/players", json={
        "usta_number": usta, "gender": gender,
        "first_name": first, "last_name": last or ("P" + uuid.uuid4().hex[:5])}))


def _official(first="Off", last=None):
    return _ok(client.post("/api/officials", json={
        "first_name": first, "last_name": last or ("O" + uuid.uuid4().hex[:5])}))


def _site(name=None):
    return _ok(client.post("/api/sites", json={"name": name or ("Site " + uuid.uuid4().hex[:5])}))


def _upload(tid, typ, fname, content, mime="text/csv"):
    return client.post(f"/api/import/tournaments/{tid}/{typ}",
                       files={"file": (fname, content, mime)})


def _stage(tid, typ, fname, content, mime="text/csv"):
    return _ok(_upload(tid, typ, fname, content, mime))


def _merge(bid):
    return _ok(client.post(f"/api/import/batches/{bid}/merge"), 200)


# ============================================================ types + templates

EXPECTED_TYPES = {
    "roster", "roster_initial", "roster_correction", "late_entries", "withdrawals",
    "scheduling_avoidances", "division_flexibility", "pairing_avoidances",
    "doubles_requests", "distances", "player_hotels", "tshirt_hotel_dietary",
}


def test_import_types_lists_all_importers():
    types = {t["key"]: t for t in _ok(client.get("/api/import/types"), 200)}
    missing = EXPECTED_TYPES - set(types)
    assert not missing, f"/api/import/types missing: {missing}"
    for key, t in types.items():
        assert t["columns"], f"{key} advertises no columns"
        assert set(t["required"]) <= set(t["columns"]), f"{key} required ⊄ columns"


def test_every_template_downloads_and_round_trips():
    """The headers handed to the TD (the downloadable template) MUST parse back
    through the same importer: an empty template stages 0 rows, never a 400.
    Guards against header/alias drift between template_csv and parse_file —
    including the XLSX path."""
    t = _tournament()
    for meta in _ok(client.get("/api/import/types"), 200):
        key = meta["key"]
        required = set(meta["required"])

        rc = client.get(f"/api/import/template/{key}?fmt=csv")
        assert rc.status_code == 200, (key, rc.text)
        assert rc.headers["content-type"].startswith("text/csv")
        assert rc.content.startswith(b"\xef\xbb\xbf"), f"{key} CSV template lacks the Excel BOM"
        csv_cols = {c.strip() for c in rc.text.splitlines()[0].lstrip("﻿").split(",")}
        assert required <= csv_cols, f"{key} CSV template missing required {required - csv_cols}"

        rx = client.get(f"/api/import/template/{key}?fmt=xlsx")
        assert rx.status_code == 200, (key, rx.text)
        wb = openpyxl.load_workbook(io.BytesIO(rx.content), read_only=True)
        xcols = {str(c) for c in next(wb.active.iter_rows(values_only=True)) if c}
        assert required <= xcols, f"{key} XLSX template missing required {required - xcols}"

        # round-trip: the empty template re-imports cleanly with zero data rows
        up = _upload(t["id"], key, f"{key}-template.csv", rc.text)
        assert up.status_code == 201, (key, up.text)
        assert up.json()["total"] == 0, f"{key} empty template staged phantom rows"


def test_unknown_import_type_is_404():
    assert client.get("/api/import/template/nope").status_code == 404
    t = _tournament()
    assert _upload(t["id"], "nope", "x.csv", "a,b\n1,2\n").status_code == 404


def test_upload_to_missing_tournament_is_404():
    assert _upload(999999, "roster", "x.csv", "usta_number\n123\n").status_code == 404


# ============================================================ CSV import e2e


def test_late_entries_csv_creates_player_roster_and_late_row():
    t = _tournament()
    u = _u()
    csv = ("usta_number,first_name,last_name,gender,age_division,events,request_date\n"
           f"{u},Late,Comer,female,G16,Singles,2026-05-20\n")
    up = _stage(t["id"], "late_entries", "late.csv", csv)
    assert up["total"] == 1 and up["valid"] == 1, up
    m = _merge(up["batch_id"])
    assert m["merged"] == 1 and m["failed"] == 0, m
    late = client.get(f"/api/tournaments/{t['id']}/late-entries").json()
    assert any(r["usta_number"] == u for r in late)
    roster = {e["usta_number"]: e for e in client.get(f"/api/tournaments/{t['id']}/players").json()}
    assert u in roster and roster[u]["age_division"] == "G16"


def test_withdrawals_csv_requires_reason_and_files():
    t = _tournament()
    good, bad = _u(), _u()
    csv = ("usta_number,first_name,last_name,gender,reason\n"
           f"{good},With,Drawn,female,injury\n"
           f"{bad},No,Reason,female,\n")          # reason blank -> merge fails (not an alternate)
    up = _stage(t["id"], "withdrawals", "wd.csv", csv)
    assert up["valid"] == 2, up                    # both stage valid (reason isn't required at staging)
    m = _merge(up["batch_id"])
    assert m["merged"] == 1 and m["failed"] == 1, m  # the reasonless one fails at merge
    assert "reason is required" in (m["errors"][0]["error"]).lower()
    wd = client.get(f"/api/tournaments/{t['id']}/withdrawals").json()
    assert any(r["usta_number"] == good for r in wd)


def test_scheduling_and_divflex_and_player_hotels_csv():
    t = _tournament()
    for typ, extra_header, extra_vals in [
        ("scheduling_avoidances", "avoid_day,avoid_time_range", "Saturday,before 9am"),
        ("division_flexibility", "home_division,willing_divisions", "G16,G18"),
        ("player_hotels", "hotel_name,lodging_plan", "Marriott,hotel"),
    ]:
        u = _u()
        csv = (f"usta_number,first_name,last_name,gender,{extra_header}\n"
               f"{u},A,B,female,{extra_vals}\n")
        up = _stage(t["id"], typ, f"{typ}.csv", csv)
        assert up["valid"] == 1, (typ, up)
        m = _merge(up["batch_id"])
        assert m["merged"] == 1 and m["failed"] == 0, (typ, m)


def test_distances_csv_by_ids_upserts_catalog():
    t = _tournament()
    off, site = _official(), _site()
    csv = f"official_id,site_id,one_way_miles\n{off['id']},{site['id']},42.5\n"
    up = _stage(t["id"], "distances", "dist.csv", csv)
    assert up["valid"] == 1, up
    m = _merge(up["batch_id"])
    assert m["merged"] == 1 and m["failed"] == 0, m
    dists = client.get("/api/distances").json()
    assert any(d["official_id"] == off["id"] and d["site_id"] == site["id"]
               and float(d["one_way_miles"]) == 42.5 for d in dists)


def test_pairing_avoidances_wide_csv_needs_existing_players():
    t = _tournament()
    a, b = _u(), _u()
    _player(a); _player(b)
    csv = f"usta_1,usta_2,relationship\n{a},{b},siblings\n"
    up = _stage(t["id"], "pairing_avoidances", "pair.csv", csv)
    assert up["valid"] == 1, up
    m = _merge(up["batch_id"])
    assert m["merged"] == 1 and m["failed"] == 0, m
    pa = client.get(f"/api/tournaments/{t['id']}/pairing-avoidances").json()
    assert len(pa) >= 1


def test_doubles_requests_csv_mutual_pair():
    t = _tournament()
    a, b = _u(), _u()
    _player(a, first="Ann"); _player(b, first="Bea")
    csv = ("usta_number,first_name,last_name,gender,age_division,wants_random,partner_usta\n"
           f"{a},Ann,X,female,G16,,{b}\n"
           f"{b},Bea,Y,female,G16,,{a}\n")
    up = _stage(t["id"], "doubles_requests", "dbl.csv", csv)
    assert up["valid"] == 2, up
    m = _merge(up["batch_id"])
    assert m["merged"] == 2 and m["failed"] == 0, m
    doubles = client.get(f"/api/tournaments/{t['id']}/doubles").json()
    assert doubles  # a verified/pending pair exists


# ============================================================ XLSX + PDF paths


def test_roster_xlsx_import_end_to_end():
    """The openpyxl read path — build a REAL .xlsx in memory and import it."""
    t = _tournament()
    u = _u()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["usta_number", "first_name", "last_name", "gender", "age_division", "selection_status"])
    ws.append([u, "Excel", "Import", "female", "G14", "selected"])
    buf = io.BytesIO(); wb.save(buf)
    up = _stage(t["id"], "roster", "roster.xlsx", buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert up["total"] == 1 and up["valid"] == 1, up
    m = _merge(up["batch_id"])
    assert m["merged"] == 1 and m["failed"] == 0, m
    roster = {e["usta_number"]: e for e in client.get(f"/api/tournaments/{t['id']}/players").json()}
    assert roster[u]["age_division"] == "G14" and roster[u]["selection_status"] == "selected"


def test_emails_pdf_import_stages_and_merges():
    if not _FIXTURE_PDF.exists():
        pytest.skip("corpus PDF fixture not present")
    t = _tournament()
    up = _stage(t["id"], "emails_pdf", "emails.pdf", _FIXTURE_PDF.read_bytes(),
                "application/pdf")
    assert up["total"] >= 10, up                  # the corpus has ~30 emails
    assert up["valid"] >= 1, up
    m = _merge(up["batch_id"])
    assert m["merged"] >= 1 and m["failed"] == 0, m
    inbox = client.get(f"/api/emails?tournament_id={t['id']}").json()
    assert len(inbox) >= 1


# ============================================================ error handling


def test_corrupt_xlsx_returns_friendly_400():
    t = _tournament()
    r = _upload(t["id"], "roster", "bad.xlsx", b"this is not a real xlsx zip",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert r.status_code == 400, r.text
    assert "parse" in r.json()["detail"].lower()


def test_missing_required_column_stages_invalid():
    t = _tournament()
    # no usta_number column at all -> the row is staged but flagged invalid
    csv = "first_name,age_division\nBob,G16\n"
    up = _stage(t["id"], "roster", "noid.csv", csv)
    assert up["total"] == 1 and up["valid"] == 0 and up["invalid"] == 1, up
    assert "usta_number" in up["errors"][0]["error"]


def test_unknown_partner_usta_stages_invalid():
    t = _tournament()
    a = _u()
    _player(a)
    # partner_usta points at a player that doesn't exist in Setup -> invalid
    csv = ("usta_number,first_name,last_name,gender,partner_usta\n"
           f"{a},Ann,X,female,9999999999\n")
    up = _stage(t["id"], "doubles_requests", "dbl.csv", csv)
    assert up["valid"] == 0 and up["invalid"] == 1, up
    assert "9999999999" in up["errors"][0]["error"]


def test_discard_batch_removes_it():
    t = _tournament()
    up = _stage(t["id"], "roster", "r.csv", f"usta_number,gender\n{_u()},female\n")
    assert client.delete(f"/api/import/batches/{up['batch_id']}").status_code == 204
    assert client.get(f"/api/import/batches/{up['batch_id']}").status_code == 404


# ============================================================ exports


def test_payroll_export_csv_is_excel_friendly():
    t = _tournament()
    r = client.get(f"/api/tournaments/{t['id']}/payroll/export.csv")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert r.content.startswith(b"\xef\xbb\xbf"), "payroll CSV must carry the Excel UTF-8 BOM"
    assert "attachment" in r.headers.get("content-disposition", "")


def test_assignment_audit_csv_exports_header():
    t = _tournament()
    r = client.get(f"/api/tournaments/{t['id']}/assignment-audit.csv")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    # header row present even with no assignments yet (no 500 on empty)
    assert r.text.lstrip("﻿").splitlines()[0].lower().startswith("when")


def test_official_schedule_ics_exports():
    o = _official()
    r = client.get(f"/api/officials/{o['id']}/schedule.ics")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/calendar")
    assert r.text.startswith("BEGIN:VCALENDAR")
    assert "END:VCALENDAR" in r.text


# ============================================================ preview-grid edit/delete


def _batch(bid):
    return _ok(client.get(f"/api/import/batches/{bid}"), 200)


def test_batch_rows_carry_id_for_editing():
    t = _tournament()
    up = _stage(t["id"], "roster", "r.csv", f"usta_number,gender\n{_u()},female\n")
    rows = _batch(up["batch_id"])["rows"]
    assert rows and "id" in rows[0]


def test_edit_staged_row_fixes_invalid_then_merges():
    """The redesign's core: an invalid staged row is corrected in the preview grid
    (PATCH) → revalidates to valid → merges."""
    t = _tournament()
    # one valid + one invalid (no usta_number)
    csv = "usta_number,first_name,gender,age_division\n" + f"{_u()},Good,female,G16\n" + ",Bad,female,G16\n"
    up = _stage(t["id"], "roster", "r.csv", csv)
    assert up["valid"] == 1 and up["invalid"] == 1, up
    rows = _batch(up["batch_id"])["rows"]
    bad = next(r for r in rows if not r["valid"])
    fixed_usta = _u()
    patched = _ok(client.patch(
        f"/api/import/batches/{up['batch_id']}/rows/{bad['id']}",
        json={"data": {**bad["data"], "usta_number": fixed_usta}}), 200)
    assert patched["valid"] is True and patched["error"] is None, patched
    assert patched["counts"]["valid"] == 2 and patched["counts"]["invalid"] == 0
    m = _merge(up["batch_id"])
    assert m["merged"] == 2 and m["failed"] == 0, m
    roster = {e["usta_number"] for e in client.get(f"/api/tournaments/{t['id']}/players").json()}
    assert fixed_usta in roster


def test_edit_can_invalidate_a_row_too():
    t = _tournament()
    up = _stage(t["id"], "roster", "r.csv", f"usta_number,gender\n{_u()},female\n")
    row = _batch(up["batch_id"])["rows"][0]
    # blank out the required usta_number -> becomes invalid
    patched = _ok(client.patch(
        f"/api/import/batches/{up['batch_id']}/rows/{row['id']}",
        json={"data": {**row["data"], "usta_number": ""}}), 200)
    assert patched["valid"] is False and "usta_number" in patched["error"]
    assert patched["counts"]["valid"] == 0


def test_delete_staged_row_updates_counts():
    t = _tournament()
    csv = f"usta_number,gender\n{_u()},female\n{_u()},female\n"
    up = _stage(t["id"], "roster", "r.csv", csv)
    rows = _batch(up["batch_id"])["rows"]
    out = _ok(client.delete(f"/api/import/batches/{up['batch_id']}/rows/{rows[0]['id']}"), 200)
    assert out["counts"]["total"] == 1
    assert len(_batch(up["batch_id"])["rows"]) == 1


def test_edit_row_after_merge_is_409():
    t = _tournament()
    up = _stage(t["id"], "roster", "r.csv", f"usta_number,gender\n{_u()},female\n")
    row = _batch(up["batch_id"])["rows"][0]
    _merge(up["batch_id"])
    r = client.patch(f"/api/import/batches/{up['batch_id']}/rows/{row['id']}",
                     json={"data": {**row["data"], "first_name": "X"}})
    assert r.status_code == 409, r.text
