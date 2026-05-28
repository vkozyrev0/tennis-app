"""Backfill officials + the official-site distance matrix from the TD's workbook.

Reads `Officials Mileage Workbook.xlsx` (repo root): columns A..E are
Name/Address/City/State/Zip and columns F/G/H are the reimbursable miles for the
JDS / RSTC / ROME sites (the sheet stores `(2*one_way) - 50`). We recover the
one-way distance as `(reimbursable + 50) / 2` and store it in
`official_site_distance` (audit §3.7).

Data-quality guards (audit §3.7 S4/S6):
  * blank cells are skipped (sparse matrix),
  * the `182` placeholder (`(2*116)-50`, reused for several officials) is skipped.

Officials are matched by (first, last) and created if missing. Sites are matched
by code; they must already exist (run seed.py first). Run from backend/:
    python backfill_distances.py
"""
from pathlib import Path

from openpyxl import load_workbook

from app.db import get_conn

WORKBOOK = Path(__file__).resolve().parent.parent / "Officials Mileage Workbook.xlsx"
COL_TO_SITE_CODE = {"F": "JDS", "G": "RSTC", "H": "ROME"}
PLACEHOLDER = 182  # (2*116)-50, reused across officials — treat as missing


def _split_name(raw: str):
    raw = (raw or "").strip()
    if "," in raw:
        last, first = raw.split(",", 1)
        return first.strip(), last.strip()
    return "", raw  # no comma -> treat whole thing as last name


def main() -> None:
    if not WORKBOOK.exists():
        print(f"workbook not found: {WORKBOOK}")
        return

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb.active

    conn = get_conn()
    created_officials = updated_officials = distances = skipped_placeholder = 0
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT code, id FROM site")
            site_id_by_code = {r["code"]: r["id"] for r in cur.fetchall()}
            missing = [c for c in COL_TO_SITE_CODE.values() if c not in site_id_by_code]
            if missing:
                print(f"missing sites {missing} — run seed.py first")
                return

            for row in range(2, ws.max_row + 1):
                name = ws[f"A{row}"].value
                if not name or not str(name).strip():
                    continue
                first, last = _split_name(str(name))
                street = ws[f"B{row}"].value
                city = ws[f"C{row}"].value
                state = ws[f"D{row}"].value
                zip_ = ws[f"E{row}"].value

                cur.execute(
                    "SELECT id FROM official WHERE lower(first_name)=lower(%s) "
                    "AND lower(last_name)=lower(%s) LIMIT 1",
                    (first, last),
                )
                found = cur.fetchone()
                if found:
                    official_id = found["id"]
                    cur.execute(
                        "UPDATE official SET street=COALESCE(%s,street), "
                        "city=COALESCE(%s,city), state=COALESCE(%s,state), "
                        "zip=COALESCE(%s,zip) WHERE id=%s",
                        (street, city, state, str(zip_) if zip_ else None, official_id),
                    )
                    updated_officials += 1
                else:
                    cur.execute(
                        "INSERT INTO official (first_name, last_name, street, city, state, zip) "
                        "VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
                        (first, last, street, city, state, str(zip_) if zip_ else None),
                    )
                    official_id = cur.fetchone()["id"]
                    created_officials += 1

                for col, code in COL_TO_SITE_CODE.items():
                    val = ws[f"{col}{row}"].value
                    if val is None:
                        continue
                    try:
                        reimbursable = float(val)
                    except (TypeError, ValueError):
                        continue
                    if int(reimbursable) == PLACEHOLDER:
                        skipped_placeholder += 1
                        continue
                    one_way = round((reimbursable + 50) / 2, 1)
                    if one_way <= 0:
                        continue
                    cur.execute(
                        """
                        INSERT INTO official_site_distance (official_id, site_id, one_way_miles, source)
                        VALUES (%s, %s, %s, 'manual')
                        ON CONFLICT (official_id, site_id)
                        DO UPDATE SET one_way_miles = EXCLUDED.one_way_miles, source = 'manual'
                        """,
                        (official_id, site_id_by_code[code], one_way),
                    )
                    distances += 1
        conn.commit()
    finally:
        conn.close()

    print(
        f"officials: +{created_officials} created, {updated_officials} updated | "
        f"distances upserted: {distances} | placeholder(182) skipped: {skipped_placeholder}"
    )


if __name__ == "__main__":
    main()
