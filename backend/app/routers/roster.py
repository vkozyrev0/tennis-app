"""Tournament <-> Player roster (tournament_entry) + CSV/XLSX import."""
import csv
import io
import re

import psycopg
from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile
from openpyxl import load_workbook

from ..db import db_dep
from ..models import RosterEntryCreate, RosterEntryOut

router = APIRouter(tags=["roster"])

# Map many header spellings (USTA exports vary) to canonical roster fields (§3.8).
_HEADER_ALIASES = {
    "usta_number": {"ustanumber", "usta", "ustano", "ustaid", "usta"},
    "first_name": {"firstname", "first", "givenname"},
    "last_name": {"lastname", "last", "surname", "familyname"},
    "age_division": {"agedivision", "division", "div", "age"},
    "events": {"events", "event"},
    "selection_status": {"selectionstatus", "status", "selection"},
    "t_shirt_size": {"tshirtsize", "tshirt", "shirt", "shirtsize", "size"},
    "dietary_preference": {"dietarypreference", "dietary", "diet", "dietaryrestrictions"},
}
_VALID_STATUS = {"selected", "alternate", "withdrawn"}


def _norm(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (h or "").strip().lower())


def _canon_headers(headers) -> dict:
    """{column_index: canonical_field} for recognised headers."""
    alias_to_canon = {}
    for canon, aliases in _HEADER_ALIASES.items():
        alias_to_canon[canon] = canon
        for a in aliases:
            alias_to_canon[a] = canon
    out = {}
    for i, h in enumerate(headers or []):
        key = alias_to_canon.get(_norm(str(h) if h is not None else ""))
        if key:
            out[i] = key
    return out


def _parse_rows(filename: str, raw: bytes) -> list[dict]:
    name = (filename or "").lower()
    records = []
    if name.endswith((".xlsx", ".xlsm")):
        ws = load_workbook(io.BytesIO(raw), data_only=True, read_only=True).active
        it = ws.iter_rows(values_only=True)
        cmap = _canon_headers(next(it, []) or [])
        for r in it:
            rec = {key: (r[i] if i < len(r) else None) for i, key in cmap.items()}
            if any(v not in (None, "") for v in rec.values()):
                records.append(rec)
    else:  # CSV / TSV
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        cmap = _canon_headers(next(reader, []))
        for r in reader:
            rec = {key: (r[i] if i < len(r) else None) for i, key in cmap.items()}
            if any(v not in (None, "") for v in rec.values()):
                records.append(rec)
    return records


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


@router.post("/api/tournaments/{tournament_id}/players/import")
async def import_roster(tournament_id: int, file: UploadFile = File(...), conn=Depends(db_dep)):
    """Upload a CSV/XLSX roster; upsert players (by USTA #) and their entries (§3.8)."""
    records = _parse_rows(file.filename, await file.read())
    created = updated = upserted = 0
    errors: list[str] = []
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM tournament WHERE id = %s", (tournament_id,))
        if cur.fetchone() is None:
            raise HTTPException(status_code=404, detail="tournament not found")
        for idx, rec in enumerate(records, start=2):  # row 1 = headers
            usta = _s(rec.get("usta_number"))
            if not usta:
                errors.append(f"row {idx}: missing USTA number")
                continue
            first, last = _s(rec.get("first_name")), _s(rec.get("last_name"))
            cur.execute("SELECT id FROM player WHERE usta_number = %s", (usta,))
            p = cur.fetchone()
            if p:
                pid = p["id"]
                if first or last:
                    cur.execute(
                        "UPDATE player SET first_name = COALESCE(%s, first_name), "
                        "last_name = COALESCE(%s, last_name) WHERE id = %s",
                        (first, last, pid),
                    )
                updated += 1
            else:
                cur.execute(
                    "INSERT INTO player (usta_number, first_name, last_name) "
                    "VALUES (%s, %s, %s) RETURNING id",
                    (usta, first, last),
                )
                pid = cur.fetchone()["id"]
                created += 1
            status = (_s(rec.get("selection_status")) or "selected").lower()
            if status not in _VALID_STATUS:
                status = "selected"
            cur.execute(
                """
                INSERT INTO tournament_entry
                    (tournament_id, player_id, age_division, events, selection_status,
                     t_shirt_size, dietary_preference, source)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 'usta_roster')
                ON CONFLICT (tournament_id, player_id) DO UPDATE SET
                    age_division = EXCLUDED.age_division, events = EXCLUDED.events,
                    selection_status = EXCLUDED.selection_status,
                    t_shirt_size = EXCLUDED.t_shirt_size,
                    dietary_preference = EXCLUDED.dietary_preference
                """,
                (tournament_id, pid, _s(rec.get("age_division")), _s(rec.get("events")),
                 status, _s(rec.get("t_shirt_size")), _s(rec.get("dietary_preference"))),
            )
            upserted += 1
    return {"created_players": created, "updated_players": updated,
            "entries": upserted, "errors": errors}

# Names are resolved POINT-IN-TIME: the version of the player's name valid as of
# the tournament's play_start_date (policy A). Falls back to the current name when
# the tournament predates any recorded version. See docs/data-model.md §PlayerHistory.
_SELECT = """
SELECT e.id, e.tournament_id, e.player_id, e.age_division, e.events,
       e.selection_status, e.t_shirt_size, e.dietary_preference,
       p.usta_number,
       COALESCE(nm.first_name, p.first_name) AS first_name,
       COALESCE(nm.last_name,  p.last_name)  AS last_name
FROM tournament_entry e
JOIN tournament t ON t.id = e.tournament_id
JOIN player p ON p.id = e.player_id
LEFT JOIN LATERAL (
    SELECT first_name, last_name
    FROM (
        SELECT first_name, last_name, valid_from
        FROM player_history WHERE player_id = e.player_id
        UNION ALL
        SELECT first_name, last_name, updated_at FROM player WHERE id = e.player_id
    ) v
    WHERE v.valid_from <= t.play_start_date::timestamptz
    ORDER BY v.valid_from DESC
    LIMIT 1
) nm ON true
"""


@router.get("/api/tournaments/{tournament_id}/players", response_model=list[RosterEntryOut])
def list_roster(tournament_id: int, conn=Depends(db_dep)):
    with conn.cursor() as cur:
        cur.execute(_SELECT + " WHERE e.tournament_id = %s ORDER BY p.last_name, p.first_name",
                    (tournament_id,))
        return cur.fetchall()


@router.post("/api/tournaments/{tournament_id}/players",
             response_model=RosterEntryOut, status_code=201)
def add_roster_entry(tournament_id: int, body: RosterEntryCreate, conn=Depends(db_dep)):
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO tournament_entry
                    (tournament_id, player_id, age_division, events,
                     selection_status, t_shirt_size, dietary_preference)
                VALUES (%(tournament_id)s, %(player_id)s, %(age_division)s, %(events)s,
                        %(selection_status)s, %(t_shirt_size)s, %(dietary_preference)s)
                RETURNING id
                """,
                {**body.model_dump(), "tournament_id": tournament_id},
            )
            new_id = cur.fetchone()["id"]
            cur.execute(_SELECT + " WHERE e.id = %s", (new_id,))
            return cur.fetchone()
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="player already on this tournament roster")
    except psycopg.errors.ForeignKeyViolation:
        raise HTTPException(status_code=400, detail="tournament_id or player_id does not exist")


@router.put("/api/roster/{entry_id}", response_model=RosterEntryOut)
def update_roster_entry(entry_id: int, body: RosterEntryCreate, conn=Depends(db_dep)):
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE tournament_entry SET
                    player_id = %(player_id)s, age_division = %(age_division)s,
                    events = %(events)s, selection_status = %(selection_status)s,
                    t_shirt_size = %(t_shirt_size)s, dietary_preference = %(dietary_preference)s
                WHERE id = %(id)s
                RETURNING id
                """,
                {**body.model_dump(), "id": entry_id},
            )
            row = cur.fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="roster entry not found")
            cur.execute(_SELECT + " WHERE e.id = %s", (entry_id,))
            return cur.fetchone()
    except psycopg.errors.UniqueViolation:
        raise HTTPException(status_code=409, detail="player already on this tournament roster")
    except psycopg.errors.ForeignKeyViolation:
        raise HTTPException(status_code=400, detail="player_id does not exist")


@router.delete("/api/roster/{entry_id}", status_code=204)
def delete_roster_entry(entry_id: int, conn=Depends(db_dep)):
    with conn.cursor() as cur:
        cur.execute("DELETE FROM tournament_entry WHERE id = %s", (entry_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="roster entry not found")
    return Response(status_code=204)
