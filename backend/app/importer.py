"""Spreadsheet import pipeline (audit §3.8).

Uploads are parsed (CSV or XLSX) into canonical rows, written to a staging area
(import_batch / import_row, migration 0020), validated per-row, then merged into
the main tables on confirm. One registry entry per importable data type.
"""
import csv
import io
import re

from openpyxl import Workbook, load_workbook

from .playerops import upsert_hotel, upsert_player

_VALID_STATUS = {"selected", "alternate", "withdrawn"}

# ---- shared parsing / normalization -----------------------------------------
def _norm(h) -> str:
    return re.sub(r"[^a-z0-9]", "", (str(h) if h is not None else "").strip().lower())


def _s(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# Shirt-size normalization (audit A51) shared with roster.py via shirtops.py.
from .shirtops import norm_shirt as _norm_shirt


# ---- column model -----------------------------------------------------------
class Col:
    def __init__(self, canon, aliases=(), required=False):
        self.canon = canon
        self.aliases = set(aliases)
        self.required = required


_PLAYER = [
    Col("usta_number", {"ustanumber", "usta", "ustano", "ustaid"}, required=True),
    Col("first_name", {"firstname", "first", "givenname"}),
    Col("last_name", {"lastname", "last", "surname", "familyname"}),
    # Gender is required when the row creates a brand-new player (audit N1).
    # Existing players keep their stored value; the column is treated as
    # optional in `cols` so existing-player imports without it still merge.
    Col("gender", {"gender", "sex", "mf"}),
]


# Audit F14: re-export the shared helper so the existing module-local name
# keeps working without a churn rewrite at every call site.
from .playerops import norm_gender as _norm_gender  # noqa: E402, F401


def _alias_map(cols):
    m = {}
    for c in cols:
        m[c.canon] = c.canon
        for a in c.aliases:
            m[a] = c.canon
    return m


def parse_file(filename: str, raw: bytes, cols) -> list[dict]:
    """Return [{row_num, data:{canon:val}}] for non-empty rows, mapping headers."""
    amap = {_norm(k): v for k, v in _alias_map(cols).items()}
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        ws = load_workbook(io.BytesIO(raw), data_only=True, read_only=True).active
        it = ws.iter_rows(values_only=True)
        headers = next(it, []) or []
        rows = list(it)
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        headers = next(reader, [])
        rows = list(reader)
    cmap = {i: amap[_norm(h)] for i, h in enumerate(headers) if _norm(str(h) if h is not None else "") in amap}
    out = []
    for n, r in enumerate(rows, start=1):
        rec = {canon: _s(r[i]) if i < len(r) else None for i, canon in cmap.items()}
        if any(v not in (None, "") for v in rec.values()):
            out.append({"row_num": n, "data": rec})
    return out


def validate(data: dict, cols, cur) -> str | None:
    """Audit F18 + fifth-pass #3: `cur` is required and every USTA-bearing
    field is pre-checked against `player` so missing-gender errors surface at
    *staging* time across wide formats (pairing_avoidances usta_1..usta_6;
    doubles_requests partner_usta) — not just the canonical usta_number."""
    missing = [c.canon for c in cols if c.required and not _s(data.get(c.canon))]
    if missing:
        return "missing " + ", ".join(missing)
    # Collect every USTA #-shaped field the row carries.
    usta_keys = ["usta_number", "partner_usta"] + [f"usta_{n}" for n in range(1, 7)]
    ustas = [_s(data.get(k)) for k in usta_keys]
    unknown = []
    for u in ustas:
        if not u:
            continue
        cur.execute("SELECT 1 FROM player WHERE usta_number = %s", (u,))
        if cur.fetchone() is None:
            unknown.append(u)
    if unknown:
        # Roster + per-row Part B importers carry their own gender column —
        # for those, an unknown canonical `usta_number` is OK *if* gender is
        # set (the merge will create the player). Every other unknown USTA #
        # (partner_usta, usta_1..usta_6 in wide formats) must already exist
        # in Setup; we don't invent genders for them.
        gender = _norm_gender(data.get("gender"))
        canonical = _s(data.get("usta_number"))
        # Drop the canonical from the unknown list when gender is set — that
        # one will be created at merge. What remains is what the TD actually
        # has to fix (audit B4).
        actionable = [u for u in unknown
                      if not (gender and u == canonical)]
        if not actionable:
            return None
        if not gender and canonical and canonical in unknown and len(unknown) == 1:
            return (f"player {canonical} isn't in Setup yet — gender column is "
                    f"required for new players (or add them via Setup → Players first)")
        return ("player(s) not in Setup yet — add them via Setup → Players first: "
                + ", ".join(actionable))
    return None


# ---- merge functions (mirror the per-resource routers) ----------------------
# Each returns a conflict note (str) when the row hit existing data, else None.
# The merge still proceeds ("go with the merge"); the note is surfaced to the user.
def _exists(cur, table, tid, pid) -> bool:
    cur.execute(f"SELECT 1 FROM {table} WHERE tournament_id = %s AND player_id = %s LIMIT 1", (tid, pid))
    return cur.fetchone() is not None


def _merge_roster(cur, tid, d):
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"), _norm_gender(d.get("gender")))
    conflict = ("already on the roster — entry overwritten"
                if _exists(cur, "tournament_entry", tid, pid) else None)
    status = (_s(d.get("selection_status")) or "selected").lower()
    if status not in _VALID_STATUS:
        status = "selected"
    cur.execute(
        """
        INSERT INTO tournament_entry
            (tournament_id, player_id, age_division, events, selection_status,
             t_shirt_size, dietary_preference, source)
        VALUES (%s,%s,%s,%s,%s,%s,%s,'usta_roster')
        ON CONFLICT (tournament_id, player_id) DO UPDATE SET
            age_division = EXCLUDED.age_division, events = EXCLUDED.events,
            selection_status = EXCLUDED.selection_status,
            t_shirt_size = EXCLUDED.t_shirt_size,
            dietary_preference = EXCLUDED.dietary_preference
        """,
        (tid, pid, _s(d.get("age_division")), _s(d.get("events")), status,
         _norm_shirt(_s(d.get("t_shirt_size"))), _s(d.get("dietary_preference"))),
    )
    return conflict


def _merge_late(cur, tid, d):
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"), _norm_gender(d.get("gender")))
    conflict = "already has a late entry — another was added" if _exists(cur, "late_entry", tid, pid) else None
    cur.execute(
        """
        INSERT INTO tournament_entry (tournament_id, player_id, age_division, events,
            selection_status, source)
        VALUES (%s,%s,%s,%s,'selected','late_entry')
        ON CONFLICT (tournament_id, player_id) DO UPDATE SET
            age_division = COALESCE(EXCLUDED.age_division, tournament_entry.age_division),
            events = COALESCE(EXCLUDED.events, tournament_entry.events)
        """,
        (tid, pid, _s(d.get("age_division")), _s(d.get("events"))),
    )
    # Audit F2: preserve source_email_id when the staged import carries it.
    cur.execute(
        "INSERT INTO late_entry (tournament_id, player_id, request_date, request_time, "
        "age_division, events, source_email_id) VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (tid, pid, _s(d.get("request_date")), _s(d.get("request_time")),
         _s(d.get("age_division")), _s(d.get("events")),
         _coerce_int(d.get("source_email_id"))),
    )
    return conflict


def _merge_withdrawal(cur, tid, d):
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"), _norm_gender(d.get("gender")))
    conflict = "already has a withdrawal — another was added" if _exists(cur, "withdrawal", tid, pid) else None
    cur.execute("SELECT selection_status FROM tournament_entry "
                "WHERE tournament_id=%s AND player_id=%s", (tid, pid))
    entry = cur.fetchone()
    was_alt = bool(entry and entry["selection_status"] == "alternate")
    reason = _s(d.get("reason"))
    if not was_alt and not reason:
        raise ValueError("a reason is required unless the player was an alternate")
    if entry:
        cur.execute("UPDATE tournament_entry SET selection_status='withdrawn' "
                    "WHERE tournament_id=%s AND player_id=%s", (tid, pid))
    cur.execute(
        "INSERT INTO withdrawal (tournament_id, player_id, events, reason, notes, was_alternate, source_email_id) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (tid, pid, _s(d.get("events")), reason, _s(d.get("notes")), was_alt,
         _coerce_int(d.get("source_email_id"))),
    )
    return conflict


def _merge_sched(cur, tid, d):
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"), _norm_gender(d.get("gender")))
    conflict = "already has a scheduling avoidance — another was added" if _exists(cur, "scheduling_avoidance", tid, pid) else None
    cur.execute(
        "INSERT INTO scheduling_avoidance (tournament_id, player_id, avoid_day, avoid_time_range, source_email_id) "
        "VALUES (%s,%s,%s,%s,%s)",
        (tid, pid, _s(d.get("avoid_day")), _s(d.get("avoid_time_range")),
         _coerce_int(d.get("source_email_id"))),
    )
    return conflict


def _merge_divflex(cur, tid, d):
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"), _norm_gender(d.get("gender")))
    conflict = "already has a division-flexibility entry — another was added" if _exists(cur, "division_flexibility", tid, pid) else None
    cur.execute(
        "INSERT INTO division_flexibility (tournament_id, player_id, home_division, willing_divisions, source_email_id) "
        "VALUES (%s,%s,%s,%s,%s)",
        (tid, pid, _s(d.get("home_division")), _s(d.get("willing_divisions")),
         _coerce_int(d.get("source_email_id"))),
    )
    return conflict


def _merge_pairing(cur, tid, d):
    """Pairing-avoidance group (audit import/export #8). Wide format: one row
    per group, with `usta_1`, `usta_2`, `usta_3?`, `usta_4?` cells. At least
    two non-blank USTAs are required. Players must already exist in Setup."""
    members = []
    for n in range(1, 7):  # support up to 6 members per group
        u = _s(d.get(f"usta_{n}"))
        if u:
            members.append(u)
    if len(members) < 2:
        raise ValueError("pairing-avoidance group needs at least 2 USTA numbers (usta_1..)")
    # Sixth-pass: pass the row's gender column (if any) for parity with the
    # other merge fns. Today validate() rejects unknown USTAs in pairing rows
    # because `cols` doesn't include gender, so this is purely defensive —
    # if `cols` ever grows a gender alias, new players become creatable.
    gender = _norm_gender(d.get("gender"))
    seen: list[int] = []
    for u in members:
        pid = upsert_player(cur, u, None, None, gender)  # 400 if not in Setup + no gender
        if pid not in seen:
            seen.append(pid)
    if len(seen) < 2:
        raise ValueError("pairing-avoidance group needs at least 2 distinct players")
    rel = _s(d.get("relationship"))
    if rel and rel not in ("same_club", "siblings"):
        rel = None
    cur.execute(
        "INSERT INTO pairing_avoidance (tournament_id, age_division, relationship, source_email_id) "
        "VALUES (%s,%s,%s,%s) RETURNING id",
        (tid, _s(d.get("age_division")), rel, _coerce_int(d.get("source_email_id"))),
    )
    gid = cur.fetchone()["id"]
    for pid in seen:
        cur.execute(
            "INSERT INTO pairing_avoidance_member (pairing_avoidance_id, player_id) VALUES (%s,%s)",
            (gid, pid),
        )
    return None


def _merge_doubles(cur, tid, d):
    """Doubles request (audit import/export #9). One row per request; mutual
    verification fires naturally when both sides land in the batch.

    Audit B3: the requester's player record is upserted here (a `gender`
    column on the row creates them; without gender, an unknown USTA fails).
    Neither the requester nor the partner is checked against the tournament
    roster — roster-membership enforcement only fires at *pair-creation*
    time (`doubles.py:_make_pair`, audit F15), not on the request itself.
    Importing requests for un-rostered players is therefore allowed; pairing
    them is not."""
    usta = _s(d.get("usta_number"))
    if not usta:
        raise ValueError("usta_number is required")
    # Sixth-pass: pass gender so a new-player row that passed staging via the
    # gender-escape-hatch in validate() can actually create the player here.
    pid = upsert_player(cur, usta, _s(d.get("first_name")), _s(d.get("last_name")), _norm_gender(d.get("gender")))
    wants_random = str(d.get("wants_random") or "").strip().lower() in ("1", "true", "yes", "y", "random")
    partner = _s(d.get("partner_usta"))
    if not wants_random and not partner:
        raise ValueError("doubles request needs a partner_usta (or set wants_random=true)")
    if partner and partner == usta:
        raise ValueError("partner_usta cannot equal the requester's own USTA #")
    division = _s(d.get("age_division"))
    if wants_random and not division:
        raise ValueError("random doubles request needs an age_division")
    conflict = None
    cur.execute(
        "SELECT 1 FROM doubles_request WHERE tournament_id = %s AND player_id = %s",
        (tid, pid),
    )
    if cur.fetchone():
        conflict = "player already has a doubles request — another was added"
    cur.execute(
        "INSERT INTO doubles_request (tournament_id, player_id, age_division, "
        "wants_random, partner_usta, source_email_id, status) "
        "VALUES (%s,%s,%s,%s,%s,%s,'pending')",
        (tid, pid, division, wants_random, partner, _coerce_int(d.get("source_email_id"))),
    )
    return conflict


def _merge_distance(cur, tid, d):
    """Setup-catalog importer: official↔site distance (audit import/export #7).

    `tid` is unused — distances are a Setup catalog, not per-tournament.
    Identifies official + site by id (preferred) or by label lookup:
      - official: (last_name, first_name)
      - site: code, then name
    Inserts new rows or updates the existing (official, site) pair.
    """
    miles = _s(d.get("one_way_miles"))
    if miles is None:
        raise ValueError("one_way_miles is required")
    try:
        miles_f = float(miles)
    except (TypeError, ValueError):
        raise ValueError(f"one_way_miles must be numeric, got {miles!r}")
    # Resolve official.
    oid = _coerce_int(d.get("official_id"))
    if oid is None:
        first, last = _s(d.get("first_name")), _s(d.get("last_name"))
        if not last:
            raise ValueError("official_id or (first_name, last_name) is required")
        # Fifth-pass #4: collect ALL matches and reject if ambiguous, so a
        # duplicate last_name doesn't get a row written against an arbitrary
        # official by storage order. Caller must use official_id or supply
        # first_name to disambiguate.
        if first:
            cur.execute(
                "SELECT id FROM official WHERE lower(last_name) = lower(%s) "
                "AND lower(first_name) = lower(%s)",
                (last, first),
            )
        else:
            cur.execute(
                "SELECT id FROM official WHERE lower(last_name) = lower(%s)",
                (last,),
            )
        rows = cur.fetchall()
        if not rows:
            raise ValueError(f"official {first or ''} {last} not found")
        if len(rows) > 1:
            raise ValueError(
                f"official {first or ''} {last} is ambiguous ({len(rows)} matches) — supply first_name or official_id"
            )
        oid = rows[0]["id"]
    # Resolve site.
    sid = _coerce_int(d.get("site_id"))
    if sid is None:
        code, name = _s(d.get("site_code")), _s(d.get("site_name"))
        if code:
            cur.execute("SELECT id FROM site WHERE code = %s", (code,))
            rows = cur.fetchall()
        elif name:
            cur.execute("SELECT id FROM site WHERE lower(name) = lower(%s)", (name,))
            rows = cur.fetchall()
        else:
            raise ValueError("site_id, site_code, or site_name is required")
        if not rows:
            raise ValueError(f"site {code or name} not found")
        if len(rows) > 1:
            raise ValueError(
                f"site {code or name} is ambiguous ({len(rows)} matches) — supply site_id"
            )
        sid = rows[0]["id"]
    raw_source = _s(d.get("source"))
    source = raw_source or "manual"
    source_note = None
    if source not in ("manual", "geocoded"):
        # Fifth-pass #7: surface the coercion to the TD rather than silently
        # downgrading. The merge still proceeds (data lands) — the conflict
        # note bubbles up in /import/batches/{id}/merge#conflicts.
        source_note = f"source {source!r} not recognized — stored as 'manual'"
        source = "manual"
    cur.execute("SELECT id FROM official_site_distance WHERE official_id = %s AND site_id = %s", (oid, sid))
    existing = cur.fetchone()
    notes = []
    if existing:
        cur.execute(
            "UPDATE official_site_distance SET one_way_miles = %s, source = %s WHERE id = %s",
            (miles_f, source, existing["id"]),
        )
        notes.append("distance already on file — overwritten")
    else:
        cur.execute(
            "INSERT INTO official_site_distance (official_id, site_id, one_way_miles, source) "
            "VALUES (%s, %s, %s, %s)",
            (oid, sid, miles_f, source),
        )
    if source_note:
        notes.append(source_note)
    return "; ".join(notes) if notes else None


def _merge_photel(cur, tid, d):
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"), _norm_gender(d.get("gender")))
    conflict = "already has a hotel on file — another was added" if _exists(cur, "player_hotel_stay", tid, pid) else None
    hid, hname = upsert_hotel(cur, d.get("hotel_name"))
    lodging = " ".join((d.get("lodging_plan") or "").split()) or None
    cur.execute(
        "INSERT INTO player_hotel_stay (tournament_id, player_id, hotel_id, hotel_name, lodging_plan, source_email_id) "
        "VALUES (%s,%s,%s,%s,%s,%s)",
        (tid, pid, hid, hname, lodging,
         _coerce_int(d.get("source_email_id"))),
    )
    return conflict


def _coerce_int(v):
    """Best-effort int coercion for staged-import values (CSV gives strings)."""
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


# ---- B2a Initial-Roster helpers --------------------------------------------
# The real "Tournament Full Player Data (June 2026).xlsx" carries data the
# existing schema and earlier importer didn't track. These helpers parse the
# real-world formats reliably and surface coercion failures as row notes
# rather than swallowing bad data.

def _coerce_decimal(v):
    """Best-effort numeric coercion for money/WTN cells. Strips $ and commas."""
    if v is None:
        return None
    s = str(v).strip().lstrip("$").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _coerce_bool(v):
    """Permissive truthiness for spreadsheet cells: 'Y'/'N', 'true'/'false',
    'yes'/'no', 'sign in'. Returns None on blanks (don't pretend to know)."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    if s in ("1", "true", "yes", "y", "sign in", "signed in", "x"):
        return True
    if s in ("0", "false", "no", "n", "absent"):
        return False
    return None


def _year_to_date(v):
    """Year-of-birth integer → 'YYYY-01-01' for storage. Caller sets
    birthdate_precision='year' so the UI can distinguish from a true DOB."""
    n = _coerce_int(v)
    if n is None or not (1900 <= n <= 2100):
        return None
    return f"{n:04d}-01-01"


# Map the verbose junior-event strings the USTA sheets use into our
# (division, event) split. The "Events" cell may hold multiple comma-separated
# entries — all for the same division for a given player. Two patterns appear
# in the real exports — the Initial sheet uses "Boys' Singles 14 & under"
# (event-then-age) and the Correction sheet uses "Girls' 14 & under singles"
# (age-then-event) — so we accept both orders.
_JUNIOR_EVENT_RES = [
    re.compile(r"(boys|girls)['’]?\s+(singles|doubles)\s+(\d+)\s*&\s*under", re.IGNORECASE),
    re.compile(r"(boys|girls)['’]?\s+(\d+)\s*&\s*under\s+(singles|doubles)", re.IGNORECASE),
]

def _parse_events_and_division(raw):
    """Split 'Boys' Singles 14 & under, Boys' Doubles 14 & under' into
    (division_code, events_str). Returns (None, raw) on no match so the
    bare-event format (already-canonical 'Singles, Doubles') still flows
    through other code paths."""
    if not raw:
        return None, None
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    division = None
    events = []
    for p in parts:
        m = None
        for rx in _JUNIOR_EVENT_RES:
            m = rx.search(p)
            if m:
                break
        if m:
            sex = m.group(1).lower()
            # First regex: groups (sex, event, age); second: (sex, age, event).
            g2, g3 = m.group(2), m.group(3)
            if g2.isdigit():
                age, evt = int(g2), g3.title()
            else:
                evt, age = g2.title(), int(g3)
            d = ("B" if sex == "boys" else "G") + str(age)
            division = division or d  # first one wins; all should match
            events.append(evt)
        else:
            # Already-canonical event name (junior "Singles"/"Doubles" or
            # adult "Men's Singles" etc.) — pass through.
            events.append(p)
    return division, ", ".join(events) if events else None


# "Selection" cell often holds e.g. "SELECTED, PRE_SELECTED" or "ALTERNATE";
# the importer needs ONE canonical value. Precedence: withdrawn > selected >
# alternate > nothing (matches what a TD would file by hand).
def _parse_selection(raw):
    if not raw:
        return "selected"
    tokens = {t.strip().lower() for t in str(raw).split(",") if t.strip()}
    if "withdrawn" in tokens:
        return "withdrawn"
    if {"selected", "pre_selected", "preselected", "main_draw"} & tokens:
        return "selected"
    if "alternate" in tokens:
        return "alternate"
    return "selected"


def _ext_player_initial(cur, pid, d):
    """B2a: after upsert_player has created/touched the row, propagate the
    extended player-catalog fields the Excel file carries. Idempotent —
    COALESCE-style so re-running doesn't blank a column the new file omits."""
    cur.execute(
        """
        UPDATE player SET
            emails           = COALESCE(%s, emails),
            phones           = COALESCE(%s, phones),
            district         = COALESCE(%s, district),
            section          = COALESCE(%s, section),
            city             = COALESCE(%s, city),
            state            = COALESCE(%s, state),
            wtn_singles      = COALESCE(%s::numeric, wtn_singles),
            wtn_singles_conf = COALESCE(%s, wtn_singles_conf),
            wtn_doubles      = COALESCE(%s::numeric, wtn_doubles),
            wtn_doubles_conf = COALESCE(%s, wtn_doubles_conf),
            birthdate        = COALESCE(%s::date, birthdate),
            birthdate_precision = CASE
                WHEN %s::date IS NOT NULL AND birthdate IS NULL THEN 'year'
                ELSE birthdate_precision END
        WHERE id = %s
        """,
        (
            _s(d.get("emails")), _s(d.get("phones")),
            _s(d.get("district")), _s(d.get("section")),
            _s(d.get("city")), _s(d.get("state")),
            _coerce_decimal(d.get("wtn_singles")), _s(d.get("wtn_singles_conf")),
            _coerce_decimal(d.get("wtn_doubles")), _s(d.get("wtn_doubles_conf")),
            _year_to_date(d.get("year_of_birth")),
            _year_to_date(d.get("year_of_birth")),
            pid,
        ),
    )


# ---- B2b Correction-import helpers -----------------------------------------
# "Draw status" cells from the USTA "Updated Status" CSV are comma-separated
# keywords describing the player's draw lifecycle ("Alternate", "Withdrawn",
# "Main draw"). Precedence: withdrawn beats selected beats alternate.
def _parse_draw_status(raw):
    if not raw:
        return None
    tokens = {t.strip().lower() for t in str(raw).split(",") if t.strip()}
    if "withdrawn" in tokens:
        return "withdrawn"
    if {"selected", "main draw", "maindraw", "main"} & tokens:
        return "selected"
    if "alternate" in tokens or "alt" in tokens:
        return "alternate"
    return None


# ---- B3 combined T-shirt + Hotel + Dietary helpers -------------------------
# The real "Tournament Players T-shirt-Hotel-Dietary" CSV asks a yes/no/local
# hotel question rather than capturing a specific hotel name. Map the common
# answers to our lodging_plan enum strings; whatever doesn't match the table
# falls back to lodging_plan_raw (column added by migration 0028) for the TD
# to triage on the player-hotels grid.
_LODGING_MAP = [
    # (substring -> canonical), checked in order (first match wins). All
    # comparisons are case-insensitive.
    ("local", "Local / family"),
    ("commuter 2", "Commuter 2+ hrs"),
    ("commuter 1", "Commuter 1-2 hrs"),
    ("commuter", "Commuter"),
    ("yes",   "Hotel"),  # "Yes, I plan to reserve..."
    ("hotel", "Hotel"),
]

def _parse_hotel_answer(raw):
    """Return (canonical_lodging_plan_or_None, raw_fallback_or_None)."""
    if not raw:
        return None, None
    s = " ".join(str(raw).split())
    low = s.lower()
    for needle, canon in _LODGING_MAP:
        if needle in low:
            return canon, None
    # Unmappable — store the raw answer for TD review.
    return None, s


# "Name" cells in this file are a single-string "First Last" (no comma);
# split into first/last for the player upsert.
def _split_name(raw):
    if not raw:
        return None, None
    parts = str(raw).strip().split(None, 1)
    if len(parts) == 1:
        return parts[0], None
    return parts[0], parts[1]


def _merge_tshirt_hotel_dietary(cur, tid, d):
    """B3: combined T-shirt + Hotel + Dietary import.

    Behavior per the 2026-05-28 questionnaire:
    - USTA # not on roster → late-add (insert new roster entry).
    - Only non-empty cells overwrite — blanks leave existing values intact.
    - Hotel free-text answer maps to lodging_plan via _LODGING_MAP;
      unmappable answers go to lodging_plan_raw for TD review."""
    # Allow either explicit first/last OR the single "Name" cell ("First Last").
    first = _s(d.get("first_name")) or _split_name(d.get("name"))[0]
    last  = _s(d.get("last_name"))  or _split_name(d.get("name"))[1]
    pid = upsert_player(cur, d["usta_number"], first, last,
                        _norm_gender(d.get("gender")))
    canon_lodging, raw_lodging = _parse_hotel_answer(d.get("hotel_answer"))
    existed = _exists(cur, "tournament_entry", tid, pid)
    conflict = "roster row updated with t-shirt/hotel/dietary" if existed else None
    cur.execute(
        """
        INSERT INTO tournament_entry
            (tournament_id, player_id, t_shirt_size, dietary_preference, source)
        VALUES (%s,%s,%s,%s,'manual')
        ON CONFLICT (tournament_id, player_id) DO UPDATE SET
            t_shirt_size       = COALESCE(EXCLUDED.t_shirt_size,
                                          tournament_entry.t_shirt_size),
            dietary_preference = COALESCE(EXCLUDED.dietary_preference,
                                          tournament_entry.dietary_preference)
        """,
        (tid, pid,
         _norm_shirt(_s(d.get("t_shirt_size"))),
         _s(d.get("dietary_preference"))),
    )
    # Lodging_plan + raw fallback land on the same row separately so the
    # COALESCE protection applies one column at a time.
    if canon_lodging or raw_lodging:
        cur.execute(
            """
            UPDATE tournament_entry SET
                lodging_plan = COALESCE(%s, lodging_plan),
                lodging_plan_raw = COALESCE(%s, lodging_plan_raw)
            WHERE tournament_id = %s AND player_id = %s
            """,
            (canon_lodging, raw_lodging, tid, pid),
        )
    return conflict


def _merge_roster_correction(cur, tid, d):
    """B2b Correction import. Applies post-withdrawal/alternate-promotion
    status changes from the USTA "Updated Status" CSV. Rules per the
    2026-05-28 questionnaire:
      - USTA # not on roster → late-add (insert with parsed status).
      - USTA # already on roster → update status + division + events
        + sign-in flag + suspension points. Other roster fields untouched.
      - Roster rows NOT in the file → untouched (this importer never deletes).
    Also propagates WTN/contact updates to Setup → Players (idempotent)."""
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"),
                        _norm_gender(d.get("gender")))
    _ext_player_initial(cur, pid, d)  # WTN / city / state / etc. flow up
    parsed_div, parsed_events = _parse_events_and_division(d.get("events"))
    division = _s(d.get("age_division")) or parsed_div
    events = parsed_events or _s(d.get("events"))
    # Status: prefer the explicit Draw-status column; otherwise fall back to
    # the canonical selection_status the template form would carry.
    status = (_parse_draw_status(d.get("draw_status"))
              or _parse_selection(d.get("selection_status")))
    signed_in = _coerce_bool(d.get("signed_in"))
    susp = _coerce_int(d.get("suspension_points"))
    existed = _exists(cur, "tournament_entry", tid, pid)
    conflict = "roster row updated" if existed else None
    # INSERT for new rows, UPDATE for existing — UPSERT with COALESCE so a
    # blank cell in the file doesn't blank an existing roster value.
    cur.execute(
        """
        INSERT INTO tournament_entry
            (tournament_id, player_id, age_division, events, selection_status,
             signed_in, suspension_points, source)
        VALUES (%s,%s,%s,%s,%s,
                -- entry_source enum doesn't distinguish Initial vs Correction;
                -- both come from the same USTA dashboard. Reuse the existing
                -- value rather than ALTER TYPE for a tag the UI doesn't read.
                COALESCE(%s::boolean, false), %s::int, 'usta_roster')
        ON CONFLICT (tournament_id, player_id) DO UPDATE SET
            age_division     = COALESCE(EXCLUDED.age_division, tournament_entry.age_division),
            events           = COALESCE(EXCLUDED.events,       tournament_entry.events),
            selection_status = COALESCE(EXCLUDED.selection_status, tournament_entry.selection_status),
            signed_in        = COALESCE(%s::boolean, tournament_entry.signed_in),
            suspension_points = COALESCE(%s::int,     tournament_entry.suspension_points)
        """,
        (tid, pid, division, events, status,
         signed_in, susp,
         signed_in, susp),
    )
    return conflict


def _merge_roster_initial(cur, tid, d):
    """B2a Initial roster import. Upserts Setup → Players (catalog) AND the
    per-tournament roster row, propagating WTN, payment, and contact info from
    the USTA "Full Player Data" export. Multi-valued "Selection" and verbose
    "Boys' Singles 14 & under" events strings are parsed into our canonical
    division + events split."""
    pid = upsert_player(cur, d["usta_number"], d.get("first_name"), d.get("last_name"),
                        _norm_gender(d.get("gender")))
    _ext_player_initial(cur, pid, d)
    # Events cell carries the division — extract it if the canonical age_division
    # column wasn't supplied separately.
    parsed_div, parsed_events = _parse_events_and_division(d.get("events"))
    division = _s(d.get("age_division")) or parsed_div
    events = parsed_events or _s(d.get("events"))
    status = _parse_selection(d.get("selection_status") or d.get("selection"))
    conflict = ("already on the roster — entry overwritten"
                if _exists(cur, "tournament_entry", tid, pid) else None)
    cur.execute(
        """
        INSERT INTO tournament_entry
            (tournament_id, player_id, age_division, events, selection_status,
             t_shirt_size, dietary_preference,
             payment_status, amount_paid, amount_refunded, amount_due,
             amount_outstanding, card_stored, source)
        -- Explicit ::numeric / ::boolean casts so psycopg can infer types
        -- when the column is omitted from the CSV (the row passes NULL and
        -- Postgres otherwise complains "could not determine data type").
        VALUES (%s,%s,%s,%s,%s,%s,%s,
                %s, %s::numeric, %s::numeric, %s::numeric, %s::numeric,
                %s::boolean, 'usta_roster')
        ON CONFLICT (tournament_id, player_id) DO UPDATE SET
            age_division = COALESCE(EXCLUDED.age_division, tournament_entry.age_division),
            events = COALESCE(EXCLUDED.events, tournament_entry.events),
            selection_status = EXCLUDED.selection_status,
            t_shirt_size = COALESCE(EXCLUDED.t_shirt_size, tournament_entry.t_shirt_size),
            dietary_preference = COALESCE(EXCLUDED.dietary_preference, tournament_entry.dietary_preference),
            payment_status = COALESCE(EXCLUDED.payment_status, tournament_entry.payment_status),
            amount_paid = COALESCE(EXCLUDED.amount_paid, tournament_entry.amount_paid),
            amount_refunded = COALESCE(EXCLUDED.amount_refunded, tournament_entry.amount_refunded),
            amount_due = COALESCE(EXCLUDED.amount_due, tournament_entry.amount_due),
            amount_outstanding = COALESCE(EXCLUDED.amount_outstanding, tournament_entry.amount_outstanding),
            card_stored = COALESCE(EXCLUDED.card_stored, tournament_entry.card_stored)
        """,
        (
            tid, pid, division, events, status,
            _norm_shirt(_s(d.get("t_shirt_size"))),
            _s(d.get("dietary_preference")),
            _s(d.get("payment_status")),
            _coerce_decimal(d.get("amount_paid")),
            _coerce_decimal(d.get("amount_refunded")),
            _coerce_decimal(d.get("amount_due")),
            _coerce_decimal(d.get("amount_outstanding")),
            _coerce_bool(d.get("card_stored")),
        ),
    )
    return conflict


# Audit F2: optional column accepted by every Part-B importer so a staged CSV
# can preserve the originating email's id when it's known. Blank = no source.
_SRC_EMAIL = Col("source_email_id", {"sourceemailid", "emailid", "source"})


# ---- registry ---------------------------------------------------------------
TYPES = {
    "roster": {"label": "Roster",
               "desc": "Players entered for the tournament (division, status, t-shirt, dietary).",
               "cols": _PLAYER + [Col("age_division", {"division", "div", "age"}),
                                  Col("events", {"event"}),
                                  Col("selection_status", {"status", "selection"}),
                                  Col("t_shirt_size", {"tshirt", "shirt", "shirtsize", "size", "tshirtsize"}),
                                  Col("dietary_preference", {"dietary", "diet", "dietaryrestrictions"})],
               "merge": _merge_roster},
    "late_entries": {"label": "Late entries",
                     "desc": "Players entering after the deadline; also added to the roster.",
                     "cols": _PLAYER + [Col("age_division", {"division", "div"}),
                                        Col("events", {"event"}),
                                        Col("request_date", {"requestdate", "date"}),
                                        Col("request_time", {"requesttime", "time"}),
                                        _SRC_EMAIL],
                     "merge": _merge_late},
    "withdrawals": {"label": "Withdrawals",
                    "desc": "Players withdrawing (reason required unless they were an alternate).",
                    "cols": _PLAYER + [Col("events", {"event"}), Col("reason"), Col("notes", {"note"}), _SRC_EMAIL],
                    "merge": _merge_withdrawal},
    "scheduling_avoidances": {"label": "Scheduling avoidances",
                              "desc": "Days/times a player cannot play (adult events).",
                              "cols": _PLAYER + [Col("avoid_day", {"avoidday", "day"}),
                                                 Col("avoid_time_range", {"avoidtime", "time", "timerange"}),
                                                 _SRC_EMAIL],
                              "merge": _merge_sched},
    "division_flexibility": {"label": "Division flexibility",
                             "desc": "Players willing to play other divisions.",
                             "cols": _PLAYER + [Col("home_division", {"homedivision", "home", "division"}),
                                                Col("willing_divisions", {"willingdivisions", "willing", "divisions"}),
                                                _SRC_EMAIL],
                             "merge": _merge_divflex},
    "pairing_avoidances": {"label": "Pairing avoidances",
                           "desc": "2+ players who must not meet in round 1 (siblings / same club). Wide format: usta_1, usta_2, [usta_3..].",
                           "cols": [
                               Col("usta_1", {"usta1"}, required=True),
                               Col("usta_2", {"usta2"}, required=True),
                               Col("usta_3", {"usta3"}),
                               Col("usta_4", {"usta4"}),
                               Col("usta_5", {"usta5"}),
                               Col("usta_6", {"usta6"}),
                               Col("age_division", {"division", "div"}),
                               Col("relationship", {"rel"}),
                               _SRC_EMAIL,
                           ],
                           "merge": _merge_pairing},
    "doubles_requests": {"label": "Doubles requests",
                         "desc": "Per-player request; mutual sides pair automatically when both rows land in the batch.",
                         "cols": _PLAYER + [Col("age_division", {"division", "div"}),
                                            Col("wants_random", {"random"}),
                                            Col("partner_usta", {"partnerusta", "partner"}),
                                            _SRC_EMAIL],
                         "merge": _merge_doubles},
    # Setup-catalog importer (audit import/export #7). Not per-tournament.
    "distances": {"label": "Distances (Setup catalog — global)",
                  "desc": "Official↔site one-way mileage. Tournament context is ignored (data is global). Match by ids OR by labels (last/first + site code/name). Ambiguous labels are rejected with a row error.",
                  "cols": [
                      Col("official_id", {"officialid"}),
                      Col("first_name", {"firstname", "first"}),
                      Col("last_name", {"lastname", "last", "surname"}),
                      Col("site_id", {"siteid"}),
                      Col("site_code", {"sitecode", "code"}),
                      Col("site_name", {"sitename", "site"}),
                      Col("one_way_miles", {"miles", "onewaymiles"}, required=True),
                      Col("source", {"src"}),
                  ],
                  "merge": _merge_distance},
    # B2a: Full Player Data import — upserts BOTH the Setup → Players catalog
    # and the per-tournament roster. Real-world aliases drawn from the USTA
    # "Tournament Full Player Data" Excel export (see docs/roadmap.md backlog).
    "roster_initial": {
        "label": "Roster — Initial (Full Player Data)",
        "desc": ("Full pre-tournament roster from the USTA \"Full Player Data\" export. "
                 "Inserts new players and updates existing ones (name, gender, "
                 "city/state, WTN, district/section, emails/phones, year of "
                 "birth), then upserts the roster row with division/events, "
                 "selection status, t-shirt, dietary, and payment snapshot. "
                 "Re-runnable: re-importing overwrites the roster, with name "
                 "and contact info merged into Setup → Players."),
        "cols": [
            Col("usta_number", {"ustanumber", "usta", "ustano", "ustaid", "id"}, required=True),
            Col("first_name", {"firstname", "first", "givenname"}),
            Col("last_name", {"lastname", "last", "surname"}),
            Col("gender", {"sex"}),
            Col("year_of_birth", {"yearofbirth", "yob", "birthyear"}),
            Col("city"),
            Col("state"),
            Col("district"),
            Col("section"),
            Col("emails", {"email"}),
            Col("phones", {"phonenumbers", "phone"}),
            Col("wtn_singles", {"wtnsingles"}),
            Col("wtn_singles_conf", {"wtnsinglesconfidence"}),
            Col("wtn_doubles", {"wtndoubles"}),
            Col("wtn_doubles_conf", {"wtndoublesconfidence"}),
            Col("age_division", {"division", "div"}),
            Col("events", {"event"}),
            Col("selection_status", {"selection", "status"}),
            Col("t_shirt_size", {"tshirt", "shirt", "shirtsize", "size", "preferredtshirtsize"}),
            Col("dietary_preference", {"dietary", "diet", "dietaryrestrictions"}),
            Col("payment_status", {"paymentstatus"}),
            Col("amount_paid", {"amountpaid"}),
            Col("amount_refunded", {"amountrefunded"}),
            Col("amount_due", {"totalamountdue", "amountdue"}),
            Col("amount_outstanding", {"amountoutstanding"}),
            Col("card_stored", {"cardstored", "card"}),
        ],
        "merge": _merge_roster_initial,
    },
    # B2b: Correction import — "Updated Status" CSV from the USTA dashboard
    # after withdrawals + alternate promotions are processed. Surgical updates
    # to selection_status, division, events, sign-in, suspension points; rows
    # not in the file are left alone; unknown USTAs are late-added.
    "roster_correction": {
        "label": "Roster — Correction (Updated Status)",
        "desc": ("Post-withdrawal status corrections from the USTA "
                 "\"Updated Status\" export. Updates selection_status, "
                 "division, events, sign-in, suspension points on existing "
                 "roster rows. Players not on the roster are late-added. "
                 "Roster rows NOT in the file are left untouched. Other "
                 "fields (t-shirt, dietary, payment) are preserved."),
        "cols": [
            Col("usta_number", {"ustanumber", "ustaid", "usta", "id"}, required=True),
            Col("first_name", {"firstname", "first"}),
            Col("last_name", {"lastname", "last", "surname"}),
            Col("gender", {"sex"}),
            Col("city"),
            Col("state"),
            Col("events", {"event"}),
            Col("age_division", {"division", "div"}),
            Col("selection_status", {"status", "selection"}),
            Col("draw_status", {"drawstatus"}),
            Col("signed_in", {"tournamentsignin", "signin", "signedin"}),
            Col("suspension_points", {"suspensionpoints"}),
            Col("wtn_singles", {"wtnsingles"}),
            Col("wtn_doubles", {"wtndoubles"}),
        ],
        "merge": _merge_roster_correction,
    },
    # B3: Combined T-shirt + Hotel + Dietary import — the post-questionnaire
    # decision: replace the three per-tab imports (t-shirt size on roster,
    # player_hotels, dietary on roster) with one upload that feeds all three
    # values in a single row per player. Only non-empty cells overwrite.
    "tshirt_hotel_dietary": {
        "label": "T-shirt + Hotel + Dietary (combined)",
        "desc": ("Single row per player carrying t-shirt size, lodging plan "
                 "(parsed from the \"Are you planning to stay overnight in a "
                 "hotel?\" question), and dietary restrictions. USTA #s not on "
                 "the roster are late-added. Blank cells leave existing roster "
                 "values intact (Initial roster's t-shirt won't be overwritten "
                 "if the new file omits the column for a row)."),
        "cols": [
            Col("usta_number", {"ustanumber", "ustaid", "uaid", "usta", "id"}, required=True),
            Col("name"),                          # "First Last" (no comma)
            Col("first_name", {"firstname", "first"}),
            Col("last_name", {"lastname", "last"}),
            Col("gender", {"sex"}),
            Col("t_shirt_size", {"tshirt", "shirt", "shirtsize", "size",
                                 "preferredtshirtsize"}),
            # Real-world column header is the literal question text — match it
            # case-insensitively + ignore non-alphanumerics (the _norm() helper
            # strips spaces/punctuation, so "Are you planning to stay overnight
            # in a hotel?" normalizes to "areyouplanningtostayovernightinahotel".
            Col("hotel_answer", {"hotelanswer", "areyouplanningtostayovernightinahotel",
                                 "hotelquestion", "lodging", "lodgingplan",
                                 "hotelplan", "hotel"}),
            Col("dietary_preference", {"dietary", "diet", "dietaryrestrictions",
                                       "dietaryrestrictionslevel2level3orlevel4"}),
        ],
        "merge": _merge_tshirt_hotel_dietary,
    },
    "player_hotels": {"label": "Player hotels",
                      "desc": "Each player's reported hotel and lodging plan.",
                      "cols": _PLAYER + [Col("hotel_name", {"hotel", "hotelname"}),
                                         Col("lodging_plan", {"lodging", "lodgingplan", "plan"}),
                                         _SRC_EMAIL],
                      "merge": _merge_photel},
}


def types_meta():
    return [{"key": k, "label": t["label"], "desc": t["desc"],
             "columns": [c.canon for c in t["cols"]],
             "required": [c.canon for c in t["cols"] if c.required]}
            for k, t in TYPES.items()]


# ---- template files ---------------------------------------------------------
def template_csv(key) -> bytes:
    headers = [c.canon for c in TYPES[key]["cols"]]
    buf = io.StringIO()
    csv.writer(buf).writerow(headers)
    return buf.getvalue().encode("utf-8-sig")


def template_xlsx(key) -> bytes:
    headers = [c.canon for c in TYPES[key]["cols"]]
    wb = Workbook()
    wb.active.append(headers)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
